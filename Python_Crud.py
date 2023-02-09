from datetime import datetime
from openpyxl import load_workbook

rut="C:\\Users\\SENA\Desktop\\Python_Excel\\Base_Crud.xlsx"
rut=r"C:\Users\SENA\Desktop\Python_Excel\Base_Crud.xlsx"

def leer(ruta:str,extraer:str):
    Archivo_Exccel = load_workbook(ruta)
    Hoja_datos = Archivo_Exccel['Datos_Crud']
    Hoja_datos = Hoja_datos['A2':'F'+str(Hoja_datos.max_row)]

    info={}

    for i in Hoja_datos:
        if isinstance(i[0].value,int):
            info.setdefault(i[0].value,{'tarea':i[1].value,'Descripcion':i[2].value,'Estado':i[3].value,'Fecha Inicio':i[4].value,
                                        'Fecha Finalizado':i[5].value})
    if not (extraer=='todo'):
        info=filtrar(info,extraer)

    for i in info:
        print('********tarea********')
        print('ID:'+ str(i)+'\n'+'Titulo: '+str(info[i]['tarea'])+'\n'+'descripcion: '+str(info[i]['descripcion'])
              +str(info[i]['Estado'])+'\n'+'Fecha Creacion:'+ str(info[i]['Fecha Inicio']+'\n'+'Fecha Finlizacion: ')
              +str(info[i]['Fecha Finalizacion']))
        print()
    return

def filtrar(info:dict,filtro:str):
    aux={}

    for i in info:
        if info[i]['Estado']==filtro:
            aux.setdefault(i,info,[i])
    return aux


def Actualizar(ruta:str,identificador:int,datos_actualizados:dict):
    Archivo_Exccel = load_workbook(ruta)
    Hoja_datos = Archivo_Exccel['Datos_Crud']
    Hoja_datos = Hoja_datos['A2':'F'+str(Hoja_datos.max_row)]
    hoja=Archivo_Exccel.active

    titulo=2
    descripcion=3
    estado=4
    Fecha_Inicio=5
    Fecha_Finalizado=6
    encontro=False
    for i in Hoja_datos:
        if i[0].value==identificador:
            fila=i[0].row
            encontro=True
            for d in datos_actualizados:
                if d=='titulo' and not(datos_actualizados[d]==''):
                    hoja.cell(row=fila, column=titulo).value=datos_actualizados[d]
                elif d=='descripcion' and not(datos_actualizados[d]==''):
                    hoja.cell(row=fila, column=descripcion).value=datos_actualizados[d]
                elif d=='estado' and not(datos_actualizados[d]==''):
                    hoja.cell(row=fila, column=estado).value=datos_actualizados[d]
                elif d=='fecha inicio' and not(datos_actualizados[d]==''):
                    hoja.cell(row=fila, column=Fecha_Inicio).value=datos_actualizados[d]
                elif d=='Fecha finalizacion' and not(datos_actualizados[d]==''):
                    hoja.cell(row=fila, column=Fecha_Finalizado).value=datos_actualizados[d]
    Archivo_Exccel.save(ruta)
    if encontro==False:
        print('Error: No existe tarea con ese Id')
        print()
    return

def agregar(ruta:int,datos:dict):
    Archivo_Exccel = load_workbook(ruta)
    Hoja_datos = Archivo_Exccel['Datos_crud']
    Hoja_datos = Hoja_datos['A2':'F'+str(Hoja_datos.max_row+1)]
    hoja=Archivo_Exccel.active

    titulo=2
    descripcion=3
    estado=4
    Fecha_Inicio=5
    Fecha_Finalizado=6
    for i in Hoja_datos:
        if not(isinstance(i[0].value,int)):
            identificador=i[0].row
            hoja.cell(row=identificador, column=1).value=identificador-1
            hoja.cell(row=identificador, column=titulo).value=datos['titulo']
            hoja.cell(row=identificador, column=descripcion).value=datos['descripcion']
            hoja.cell(row=identificador, column=estado).value=datos['estado']
            hoja.cell(row=identificador, column=Fecha_Inicio).value=datos['fecha inicio']
            hoja.cell(row=identificador, column=Fecha_Finalizado).value=datos['fecha finalizacion']
            break
    Archivo_Exccel.save(ruta)
    return


def borrar(ruta,identificador):
    Archivo_Exccel = load_workbook(ruta)
    Hoja_datos = Archivo_Exccel['Datos_Crud']
    Hoja_datos = Hoja_datos['A2':'F'+str(Hoja_datos.max_row)]
    hoja=Archivo_Exccel.active

    titulo=2
    descripcion=3
    estado=4
    Fecha_Inicio=5
    Fecha_Finalizado=6
    encontro=False
    for i in Hoja_datos:
        if i[0].value==identificador:
            fila=i[0].row
            encontro=True

            hoja.cell(row=fila, column=1).value=""
            hoja.cell(row=fila, column=titulo).value=""
            hoja.cell(row=fila, column=descripcion).value=""
            hoja.cell(row=fila, column=estado).value=""
            hoja.cell(row=fila, column=Fecha_Inicio).value=""
            hoja.cell(row=fila, column=Fecha_Finalizado).value=""
    Archivo_Exccel.save(ruta)
    if encontro==False:
        print('Error: No existe una tarea con ese id')
        print()
    return

rut="C:\\Users\\SENA\Desktop\\Python_Excel\\Base_Crud.xlsx"

datos_actualizados={'titulo':'','descripcion':'','estado':'','Fecha Inicio':'','Fecha Finalizacion':''}
while True:
    print('Indique la accion que desea realizar: ')
    print('Consultar: 1')
    print('Actualizar: 2')
    print('Crear Nueva Tarea: 3')
    print('Borrar: 4')
    accion=input('Escriba la opcion')

    if not (accion=='1') and not (accion=='2') and not(accion=='3') and not(accion=='4'):
        print('Comando invalido por favor elija una opcion valida')
    elif accion=='1':
        opc_consulta=''
        print('Indique la tarea que desea consultar: ')
        print('todas las tareas: 1')
        print('En espera: 2')
        print('En ejecucion: 3')
        print('Por Aprobar')
        print('Finalizado: 5')
        opc_consulta=input('Escriba la tarea que desea consultar: ')
        if opc_consulta=='1':
            print()
            print()
            print('*** Consultando todas las tareas **')
            leer(rut,'todo')
        elif opc_consulta=='2':
            print()
            print()
            print('*** Consultando tareas en espera')
            leer(rut,'en espera')
        elif opc_consulta=='3':
            print()
            print()
            print('*** Consultando tareas en ejecucion **')
            leer(rut,'por aprobar')
        elif opc_consulta=='4':
            print()
            print()
            print('*** Consultando tareas por acbar **')
            leer(rut,'por aprobar')
        elif opc_consulta=='5':
            print()
            print()
            print('*** Consultando tareas finalizadas **')
            leer(rut,'finalizada')
    elif accion=='2':
        datos_actualizados={'titulo':'','descripcion':'','estado':'','Fecha inicio':'','Fecha finalizacion':''}
        print('** Actualizar tarea **')
        print()
        id_Actualizar=int(input('Indique el Id de la tarea que desea actualzar: '))
        print()
        print('** Nuevo titulo **')
        print('** Nota: si no desea actualizar el titulo solo oprima ENTER')
        datos_actualizados['titulo']=input('Indique el nuevo titulo de la tarea')
        print()
        print('** Nueva descripcion **')
        print('** Nota: si no desea actualizar la descripcion solo oprima ENTER **')
        datos_actualizados['descripcion']=input('Indique la nueva descripcion de la tarea')
        print()
        print('** Nuevo estado **')
        print('en espera: 2')
        print('En ejecucion: 3')
        print('Por Aprobar: 4')
        print('Finalizada: 5')
        print('** Nota: si no desea actualizar el estado solo oprima ENTER')
        estado_nuevo= input('Indique el nuevo estado de la tarea: ')
        if estado_nuevo=='2':
            datos_actualizados['estado']='En espera'
        elif estado_nuevo=='3':
            datos_actualizados['estado']='En ejecucion'
        elif estado_nuevo=='4':
            datos_actualizados['estado']='Por Aprobar'